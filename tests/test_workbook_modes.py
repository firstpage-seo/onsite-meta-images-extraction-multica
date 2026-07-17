import importlib.util
import os
import tempfile
import unittest

import openpyxl
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as SpreadsheetImage
from PIL import Image as PillowImage


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "scripts", "onsite_checklist.py")
SPEC = importlib.util.spec_from_file_location("onsite_checklist", SCRIPT)
onsite = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(onsite)


def empty_buckets():
    return {key: [] for key in onsite.TABS}


def checklist_row(ws, item):
    for row in range(3, ws.max_row + 1):
        try:
            key = f"{round(float(ws.cell(row, 2).value), 2):.1f}"
        except (TypeError, ValueError):
            continue
        if key == item:
            return row
    raise AssertionError(f"Checklist item not found: {item}")


class WorkbookModesTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.seo = onsite.default_template("seo")
        self.geo = onsite.default_template("seo_geo")
        self.meta = {"images_crawled": True, "images_sizes_known": True}

    def tearDown(self):
        self.tmp.cleanup()

    def path(self, name):
        return os.path.join(self.tmp.name, name)

    def test_detects_both_template_profiles(self):
        self.assertEqual(onsite.detect_profile(self.seo), "seo")
        self.assertEqual(onsite.detect_profile(self.geo), "seo_geo")

    def test_review_merges_rows_and_preserves_human_work(self):
        previous_buckets = empty_buckets()
        previous_buckets["17.1"] = [
            {"Address": "https://example.com/persistent"},
            {"Address": "https://example.com/resolved"},
        ]
        previous_buckets["20.1"] = [
            {"Source": "https://example.com/persistent", "Destination": "https://cdn.example/a.png"},
            {"Source": "https://example.com/resolved", "Destination": "https://cdn.example/b.png"},
        ]
        previous = self.path("previous.xlsx")
        onsite.build(self.seo, previous, previous_buckets, self.meta, mode="begin")

        wb = openpyxl.load_workbook(previous)
        missing = wb["17.1 Page titles - Missing"]
        missing.cell(4, 2).value = "Keep proposed title"
        missing.cell(4, 3).value = "Keep remark"
        missing.cell(4, 3).hyperlink = "https://notes.example/"
        missing.cell(4, 3).comment = Comment("Keep comment", "SEO")
        wb["17.2 Page titles - Duplicate"].sheet_state = "hidden"
        png = self.path("screenshot.png")
        PillowImage.new("RGB", (4, 4), "red").save(png)
        alt_sheet = wb["20.1 Image - Alt Text Missing"]
        alt_sheet.add_image(SpreadsheetImage(png), "C4")
        alt_sheet.add_image(SpreadsheetImage(png), "C5")
        checklist = wb[onsite.CHECK_TAB]
        unrelated = checklist_row(checklist, "10.1")
        checklist.cell(unrelated, onsite.CHECK_COL).value = "✖"
        wb.save(previous)

        current_buckets = empty_buckets()
        current_buckets["17.1"] = [
            {"Address": "https://example.com/persistent"},
            {"Address": "https://example.com/new"},
        ]
        current_buckets["17.2"] = [
            {"Address": "https://example.com/duplicate", "Title 1": "Duplicate title"}
        ]
        current_buckets["17.4"] = [
            {"Address": "https://example.com/multiple", "Title 1": "One", "Title 2": "Two"}
        ]
        current_buckets["20.1"] = [
            {"Source": "https://example.com/persistent", "Destination": "https://cdn.example/a.png"}
        ]

        refreshed = self.path("refreshed.xlsx")
        _, changes = onsite.build(
            previous, refreshed, current_buckets, self.meta, mode="review",
            blank_unverified=False,
        )

        wb = openpyxl.load_workbook(refreshed)
        missing = wb["17.1 Page titles - Missing"]
        self.assertEqual(missing.cell(4, 2).value, "Keep proposed title")
        self.assertEqual(missing.cell(4, 3).value, "Keep remark")
        self.assertEqual(missing.cell(4, 3).hyperlink.target, "https://notes.example/")
        self.assertEqual(missing.cell(4, 3).comment.text, "Keep comment")
        self.assertEqual(missing.cell(5, 1).value, "https://example.com/new")
        self.assertIsNone(missing.cell(6, 1).value)
        self.assertEqual(wb["17.2 Page titles - Duplicate"].sheet_state, "visible")
        self.assertEqual(wb["17.4 Page Titles - Multiple"].cell(4, 2).value, "One")
        self.assertEqual(wb["17.4 Page Titles - Multiple"].cell(4, 3).value, "Two")
        alt_sheet = wb["20.1 Image - Alt Text Missing"]
        self.assertEqual(len(alt_sheet._images), 1)
        self.assertEqual(onsite._image_position(alt_sheet._images[0]), (4, 3))
        self.assertEqual(wb[onsite.CHECK_TAB].cell(unrelated, onsite.CHECK_COL).value, "✖")
        self.assertEqual(changes["17.1"], {"new": 1, "persistent": 1, "resolved": 1})
        self.assertEqual(changes["17.2"], {"new": 1, "persistent": 0, "resolved": 0})
        self.assertEqual(changes["20.1"], {"new": 0, "persistent": 1, "resolved": 1})

    def test_geo_20_3_is_blank_for_begin_and_unchanged_for_review(self):
        begun = self.path("geo-begin.xlsx")
        onsite.build(self.geo, begun, empty_buckets(), self.meta, mode="begin")
        wb = openpyxl.load_workbook(begun)
        checklist = wb[onsite.CHECK_TAB]
        row = checklist_row(checklist, "20.3")
        self.assertIsNone(checklist.cell(row, onsite.CHECK_COL).value)
        checklist.cell(row, onsite.CHECK_COL).value = "✖"
        wb.save(begun)

        reviewed = self.path("geo-review.xlsx")
        onsite.build(
            begun, reviewed, empty_buckets(), self.meta, mode="review",
            blank_unverified=False,
        )
        wb = openpyxl.load_workbook(reviewed)
        self.assertEqual(wb[onsite.CHECK_TAB].cell(row, onsite.CHECK_COL).value, "✖")

    def test_unknown_image_evidence_writes_question_mark(self):
        output = self.path("unknown.xlsx")
        onsite.build(
            self.seo, output, empty_buckets(),
            {"images_crawled": False, "images_sizes_known": False}, mode="begin",
        )
        wb = openpyxl.load_workbook(output)
        checklist = wb[onsite.CHECK_TAB]
        self.assertEqual(checklist.cell(checklist_row(checklist, "20.1"), onsite.CHECK_COL).value, "?")
        self.assertEqual(checklist.cell(checklist_row(checklist, "20.2"), onsite.CHECK_COL).value, "?")


if __name__ == "__main__":
    unittest.main()
