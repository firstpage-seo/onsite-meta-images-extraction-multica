#!/usr/bin/env python3
"""
Get crawl data from Screaming Frog and produce either the SEO or SEO/GEO
FirstPage Onsite Checklist deliverable for items 17.1-20.2.

Two phases, independently runnable:
  crawl  -> Screaming Frog CLI writes CSV exports to a folder
  build  -> read those CSVs, apply house thresholds, write the .xlsx

Use --exports-dir to rebuild from an existing export folder without re-crawling.
"""

import argparse
import collections
import copy
import csv
import datetime
import glob
import hashlib
import os
import shutil
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

SF_BIN = "/Applications/Screaming Frog SEO Spider.app/Contents/MacOS/ScreamingFrogSEOSpiderLauncher"

# ---- House thresholds. These live here, NOT in Screaming Frog's config, so a
# ---- stale GUI setting can never silently change the numbers.
TITLE_PX = 580          # 17.3  (per checklist tab header)
DESC_PX = 880           # 18.3  (per tab note: Config > Spider > Preferences > 880 mobile max)
IMG_KB = 200            # 20.2
KB = 1024               # SF's kB convention is unconfirmed (1000 vs 1024); 100kB export's
                        # smallest row was 102876 bytes, which exceeds both. Affects only
                        # images within ~4.8KB of the cutoff.

EXPORT_TABS = [
    "Page Titles:All", "Page Titles:Missing", "Page Titles:Duplicate", "Page Titles:Multiple",
    "Meta Description:All", "Meta Description:Missing", "Meta Description:Duplicate",
    "H1:All", "H1:Missing", "H1:Duplicate", "H1:Multiple",
    "Images:All",
]
BULK_EXPORTS = [
    "Images:Images Missing Alt Attribute & Text Inlinks",
    "Images:All Image Inlinks",
]


# --------------------------------------------------------------------------- crawl

def run_crawl(url, sitemap, outdir, config=None, client=None, date=None, crawl_file=None):
    outdir = os.path.abspath(outdir)          # MUST be absolute: the launcher resets cwd to
    os.makedirs(outdir, exist_ok=True)        # the app bundle, so relative paths hit /Applications
    if not os.path.exists(SF_BIN):
        sys.exit(f"Screaming Frog not found at {SF_BIN}")

    cmd = [SF_BIN, "--headless"]

    if crawl_file:
        # Export from a crawl the user already ran in the GUI. Their settings, their crawl
        # analysis, no config file needed. Loading from a FILE path works even when a crawl
        # is open in the GUI; loading by database ID does not.
        cmd += ["--load-crawl", os.path.abspath(crawl_file)]
        print(f"[load]  crawl <- {crawl_file}")
    else:
        # Always start a Spider crawl. --crawl-sitemap opens the saved crawl in List mode.
        # A client sitemap cannot be injected into "Crawl These Sitemaps" through the CLI,
        # so the local headless route deliberately accepts website-only crawls.
        if not url:
            sys.exit("The local headless route needs a website URL.")
        cmd += ["--crawl", url]
        if config:
            cmd += ["--config", os.path.abspath(config)]
        # --save-crawl puts the crawl in SF's DB so it can be opened in the GUI afterwards.
        # Name it, or it lands in the crawl list as a bare URL and is a pain to find later.
        if client:
            cmd += ["--task-name", f"{client} - {date}" if date else client,
                    "--project-name", client]
        cmd += ["--save-crawl"]
        print(f"[crawl] Spider mode from {url}")
        print(f"[crawl] config  -> {config or 'Screaming Frog DEFAULTS (not your GUI settings)'}")

    cmd += [
        "--output-folder", outdir,
        "--export-format", "csv",
        "--overwrite",
        "--export-tabs", ",".join(EXPORT_TABS),
        "--bulk-export", ",".join(BULK_EXPORTS),
    ]
    print(f"[run]   output -> {outdir}")
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        print(proc.stdout[-3000:])
        print(proc.stderr[-3000:], file=sys.stderr)
        sys.exit(f"Screaming Frog exited {proc.returncode}")
    for line in proc.stdout.splitlines():
        if "Completed the spider" in line or "completed urls through SpiderResults" in line:
            print("[run]   " + line.split("INFO  - ")[-1])
    return outdir


# --------------------------------------------------------------------------- io

def find_csv(outdir, *must, exclude=()):
    """Locate an export by keyword. Filenames are SF's own slugs, so match loosely."""
    for path in sorted(glob.glob(os.path.join(outdir, "*.csv"))):
        name = os.path.basename(path).lower()
        if all(m in name for m in must) and not any(x in name for x in exclude):
            return path
    return None


def load(path):
    if not path or not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def col(row, *names):
    """Fetch the first present column; SF header names vary between exports."""
    for n in names:
        if n in row and row[n] not in (None, ""):
            return row[n]
    return None


def num(v):
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0


def indexable(rows):
    return [r for r in rows if (r.get("Indexability") or "").strip() == "Indexable"]


def group_by(rows, *keys):
    """Order duplicate rows so identical values sit together, biggest group first."""
    val = lambda r: (col(r, *keys) or "").strip()
    counts = collections.Counter(val(r) for r in rows)
    return sorted(rows, key=lambda r: (-counts[val(r)], val(r), col(r, "Address") or ""))


# --------------------------------------------------------------------------- buckets

def build_buckets(d):
    t_all = load(find_csv(d, "page_titles", "all"))
    d_all = load(find_csv(d, "meta_description", "all"))
    i_all = load(find_csv(d, "images_all"))

    # A failed crawl still writes header-only CSVs, which would sail through as
    # "zero issues everywhere" and produce an all-ticks deliverable for a site
    # nobody actually crawled. Refuse rather than report a false clean bill.
    if not t_all:
        sys.exit("No pages in Page Titles:All - the crawl produced nothing. "
                 "Refusing to build a deliverable that would read as an all-clear.")

    # 19.2 = cross-page H1-1 duplicates. Start from SF's Duplicate filter (which already
    # excludes noindex pages, handles pagination, etc. - do NOT reproduce that by grouping
    # h1_all, which pulls in noindex pages SF omits and inflates the count).
    #
    # SF's Duplicate ALSO flags pages whose H1-1 == H1-2 on the SAME page. Drop those - but
    # ONLY when that H1-1 value is unique within the set. A page can both repeat its H1 on
    # the page AND share that H1-1 with other pages (crefit /clearcardcredit/, H1-1
    # "健康貸度 由你掌控" on 4 pages); that is a real cross-page duplicate and must stay. The
    # earlier "drop every H1-1==H1-2 row" rule wrongly removed those. Pure same-page
    # collisions (unique H1-1) belong in 19.3 Multiple, where they already appear.
    h_dupe_raw = load(find_csv(d, "h1", "duplicate"))
    hnorm = lambda r: (col(r, "H1-1") or "").strip().lower()
    dup_h1_counts = collections.Counter(hnorm(r) for r in h_dupe_raw if hnorm(r))
    h_dupe = [r for r in h_dupe_raw
              if (col(r, "H1-1") or "").strip() != (col(r, "H1-2") or "").strip()
              or dup_h1_counts[hnorm(r)] > 1]

    # 20.2: read Size (Bytes) straight from All Image Inlinks - it carries a size on every
    # row, INCLUDING external-domain images. (An earlier version joined to Images:All for
    # size, but Images:All is internal-only, so it silently dropped every external-CDN image
    # - Prismic, Cloudinary, agentpro, etc. Reading size from the inlink row fixes that.)
    # Threshold stays in code (IMG_KB), not SF's configured "Over X kB".
    inlinks = load(find_csv(d, "all_image_inlinks"))
    alt_missing = load(find_csv(d, "missing_alt_attribute"))
    i_over = []
    sizes_seen = False
    for r in inlinks:
        size = num(col(r, "Size (Bytes)", "Size"))
        if size > 0:
            sizes_seen = True
        if size > IMG_KB * KB:
            i_over.append({"Source": col(r, "Source"),
                           "Destination": col(r, "Destination"), "Size (Bytes)": int(size)})
    i_over.sort(key=lambda x: -x["Size (Bytes)"])

    buckets = {
        "17.1": load(find_csv(d, "page_titles", "missing")),
        "17.2": group_by(load(find_csv(d, "page_titles", "duplicate")), "Title 1"),
        "17.3": sorted([r for r in indexable(t_all) if num(col(r, "Title 1 Pixel Width")) > TITLE_PX],
                       key=lambda r: -num(col(r, "Title 1 Pixel Width"))),
        "17.4": load(find_csv(d, "page_titles", "multiple")),
        "18.1": load(find_csv(d, "meta_description", "missing")),
        "18.2": group_by(load(find_csv(d, "meta_description", "duplicate")), "Meta Description 1"),
        "18.3": sorted([r for r in indexable(d_all) if num(col(r, "Meta Description 1 Pixel Width")) > DESC_PX],
                       key=lambda r: -num(col(r, "Meta Description 1 Pixel Width"))),
        "19.1": load(find_csv(d, "h1", "missing")),
        "19.2": group_by(h_dupe, "H1-1"),
        "19.3": load(find_csv(d, "h1", "multiple")),
        "20.1": alt_missing,
        "20.2": i_over,
    }
    # Did the crawl capture ANY images? Use every image export, not just Images:All -
    # Images:All lists INTERNAL images only, so a site on an external CDN has an empty
    # Images:All yet real image issues in the inlinks/alt exports (crefit: 0 in Images:All,
    # 17 missing-alt on prismic.io). If NOTHING across the image exports has rows, the crawl
    # genuinely captured no images (image crawling off, or JS rendering off on a JS-loaded
    # site) - flag 20.1/20.2 as UNKNOWN rather than report a false all-clear.
    #
    # images_sizes_known says whether All Image Inlinks supplied any byte sizes at all. If
    # images exist but that export has no sizes, a zero oversize result is unknown, not clean.
    meta = {
        "images_crawled": bool(i_all or inlinks or alt_missing),
        "images_sizes_known": sizes_seen,
    }
    return buckets, meta


# --------------------------------------------------------------------------- workbook

s = lambda r, *k: ((col(r, *k) or "").strip() or None)

# Machine columns are refreshed from the crawl. Preserve columns contain human work that is
# carried forward for persistent issues in Review mode. key_cols identify an issue across runs.
TABS = {
    "17.1": {"tab": "17.1 Page titles - Missing",
             "rowfn": lambda r: [col(r, "Address")], "machine_cols": 1,
             "preserve_cols": (2, 3), "key_cols": (1,)},
    "17.2": {"tab": "17.2 Page titles - Duplicate",
             "rowfn": lambda r: [col(r, "Address"), s(r, "Title 1")], "machine_cols": 2,
             "preserve_cols": (3, 4), "key_cols": (1, 2)},
    "17.3": {"tab": "17.3 Page Titles - Too Long",
             "rowfn": lambda r: [col(r, "Address"), s(r, "Title 1"),
                                   int(num(col(r, "Title 1 Pixel Width")))], "machine_cols": 3,
             "preserve_cols": (4,), "key_cols": (1,)},
    "17.4": {"tab": "17.4 Page Titles - Multiple",
             "rowfn": lambda r: [col(r, "Address"), s(r, "Title 1"), s(r, "Title 2")],
             "machine_cols": 3, "preserve_cols": (4,), "key_cols": (1,)},
    "18.1": {"tab": "18. Descriptions - Missing",
             "rowfn": lambda r: [col(r, "Address")], "machine_cols": 1,
             "preserve_cols": (2,), "key_cols": (1,)},
    "18.2": {"tab": "18.2 Descriptions - Duplicate",
             "rowfn": lambda r: [col(r, "Address"), s(r, "Meta Description 1")],
             "machine_cols": 2, "preserve_cols": (), "key_cols": (1, 2)},
    "18.3": {"tab": "18.3 Descriptions - Too Long",
             "rowfn": lambda r: [col(r, "Address"), s(r, "Meta Description 1"),
                                   int(num(col(r, "Meta Description 1 Pixel Width")))],
             "machine_cols": 3, "preserve_cols": (), "key_cols": (1,)},
    "19.1": {"tab": "19.1 H1 - Missing",
             "rowfn": lambda r: [col(r, "Address")], "machine_cols": 1,
             "preserve_cols": (2,), "key_cols": (1,)},
    "19.2": {"tab": "19.2 H1 - Duplicate",
             "rowfn": lambda r: [col(r, "Address"), s(r, "H1-1")], "machine_cols": 2,
             "preserve_cols": (3,), "key_cols": (1, 2)},
    "19.3": {"tab": "19.3 H1 - Multiple",
             "rowfn": lambda r: [col(r, "Address"), s(r, "H1-1"), s(r, "H1-2")],
             "machine_cols": 3, "preserve_cols": (), "key_cols": (1,)},
    "20.1": {"tab": "20.1 Image - Alt Text Missing",
             "rowfn": lambda r: [col(r, "Source"), col(r, "Destination")], "machine_cols": 2,
             "preserve_cols": (3,), "key_cols": (1, 2)},
    "20.2": {"tab": "20.2 Image - Oversize",
             "rowfn": lambda r: [r["Source"], r["Destination"], r["Size (Bytes)"]],
             "machine_cols": 3, "preserve_cols": (), "key_cols": (1, 2)},
}

PROFILES = {
    "seo": {
        "template": "Onsite Checklist - TEMPLATE - 2026.xlsx",
        "output": "Onsite Checklist - {client} - {date}.xlsx",
        "url": "https://raw.githubusercontent.com/firstpage-seo/onsite-meta-images-extraction-multica/main/template/Onsite%20Checklist%20-%20TEMPLATE%20-%202026.xlsx",
        "sha256": "69a5a6b0fb78602cbbf98093379d4d4fe0906c31eecb5f33f567d9e668eb4cd6",
    },
    "seo_geo": {
        "template": "[Working Draft_Internal SEO_GEO] Onsite Checklist - [TEMPLATE] - 20260714.xlsx",
        "output": "SEO_GEO Onsite Checklist - {client} - {date}.xlsx",
        "url": "https://raw.githubusercontent.com/firstpage-seo/onsite-meta-images-extraction-multica/main/template/%5BWorking%20Draft_Internal%20SEO_GEO%5D%20Onsite%20Checklist%20-%20%5BTEMPLATE%5D%20-%2020260714.xlsx",
        "sha256": "4efcf9f3a6ea08b423e192f727049cf07a86649bb30659df1c10905d5d68ce12",
    },
}

APPROVED_GITHUB_OWNER = "firstpage-seo"
APPROVED_GITHUB_REPO = "onsite-meta-images-extraction-multica"
MAX_TEMPLATE_BYTES = 5 * 1024 * 1024

DATA_ROW = 4          # rows 1-2 metadata, row 3 headers, data from row 4
CHECK_TAB = "SEO Implementation Checklist"
CHECK_COL = 4         # column D, "Initial Check"


def find_check_rows(ws):
    """Map '17.1' -> sheet row. Located by column B value, not hardcoded, because
    row positions shift whenever the template gains items."""
    out = {}
    for r in range(3, ws.max_row + 1):
        b = ws.cell(r, 2).value
        try:
            key = f"{round(float(b), 2):.1f}"     # col B carries float noise: 17.400000000000006
        except (TypeError, ValueError):
            continue
        if key in TABS and key not in out:
            out[key] = r
    return out


def _norm_key(value):
    return str(value or "").strip()


def _issue_key(values, key_cols):
    return tuple(_norm_key(values[c - 1]) for c in key_cols)


def _change_counts(previous, current):
    old = collections.Counter(previous)
    new = collections.Counter(current)
    persistent = sum((old & new).values())
    return {
        "new": sum((new - old).values()),
        "persistent": persistent,
        "resolved": sum((old - new).values()),
    }


def _image_position(image):
    """Return a loaded image's one-based anchor row/column when available."""
    anchor = image.anchor
    if isinstance(anchor, str):
        cell = openpyxl.utils.cell.coordinate_from_string(anchor)
        return cell[1], openpyxl.utils.cell.column_index_from_string(cell[0])
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return marker.row + 1, marker.col + 1


def _capture_cell(cell):
    return cell.value, copy.copy(cell.hyperlink), copy.copy(cell.comment)


def _restore_cell(cell, captured):
    value, hyperlink, comment = captured
    cell.value = value
    cell.hyperlink = hyperlink
    cell.comment = comment


def detect_profile(workbook_path):
    """Identify the template family from workbook structure, never its filename alone."""
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exc:
        sys.exit(f"Could not open base workbook as XLSX: {exc}")
    sheets = set(wb.sheetnames)
    wb.close()
    required = {CHECK_TAB, *(spec["tab"] for spec in TABS.values())}
    missing = sorted(required - sheets)
    if missing:
        sys.exit(f"Workbook is missing required sheets: {missing}")
    return "seo_geo" if "20.3 Image - Next-gen" in sheets else "seo"


def normalize_template_url(url):
    """Convert an approved GitHub file page to its raw URL and reject other repositories."""
    parsed = urllib.parse.urlsplit(url.strip())
    if parsed.scheme.lower() != "https":
        raise ValueError("Template URL must use HTTPS")

    host = (parsed.hostname or "").lower()
    parts = [part for part in parsed.path.split("/") if part]
    owner_repo = [APPROVED_GITHUB_OWNER, APPROVED_GITHUB_REPO]

    if host == "github.com":
        if len(parts) < 5 or parts[:2] != owner_repo or parts[2] != "blob":
            raise ValueError("Template URL must point to a file in the approved FirstPage repository")
        ref = parts[3]
        file_path = urllib.parse.quote(urllib.parse.unquote("/".join(parts[4:])), safe="/")
        return f"https://raw.githubusercontent.com/{owner_repo[0]}/{owner_repo[1]}/{ref}/{file_path}"

    if host == "raw.githubusercontent.com":
        if len(parts) < 4 or parts[:2] != owner_repo:
            raise ValueError("Template URL must point to the approved FirstPage repository")
        raw_path = urllib.parse.quote(urllib.parse.unquote("/".join(parts)), safe="/")
        return urllib.parse.urlunsplit(("https", host, "/" + raw_path, "", ""))

    raise ValueError("Template URL host must be github.com or raw.githubusercontent.com")


def download_template(url, checklist_type, output_dir, expected_sha256=None):
    """Download and validate an approved template into a temporary task directory."""
    try:
        raw_url = normalize_template_url(url)
    except ValueError as exc:
        sys.exit(f"Invalid template URL: {exc}")

    request = urllib.request.Request(raw_url, headers={"User-Agent": "FirstPage-Onsite-Skill/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            length = response.headers.get("Content-Length")
            if length and int(length) > MAX_TEMPLATE_BYTES:
                sys.exit(f"Template download exceeds {MAX_TEMPLATE_BYTES // (1024 * 1024)} MB")
            chunks = []
            total = 0
            while True:
                chunk = response.read(64 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > MAX_TEMPLATE_BYTES:
                    sys.exit(f"Template download exceeds {MAX_TEMPLATE_BYTES // (1024 * 1024)} MB")
                chunks.append(chunk)
    except (urllib.error.URLError, TimeoutError, OSError, ValueError) as exc:
        sys.exit(f"Could not download template: {exc}")

    payload = b"".join(chunks)
    if not payload.startswith(b"PK"):
        sys.exit("Downloaded template is not an XLSX file (ZIP signature missing)")

    digest = hashlib.sha256(payload).hexdigest()
    if expected_sha256 and digest != expected_sha256:
        sys.exit("Downloaded template checksum does not match the approved version")

    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, PROFILES[checklist_type]["template"])
    with open(path, "wb") as fh:
        fh.write(payload)
    print(f"[template] downloaded {checklist_type} <- {raw_url}")
    return path


def resolve_begin_template(local_path, template_url, checklist_type, output_dir):
    if local_path:
        return local_path
    source = template_url or PROFILES[checklist_type]["url"]
    try:
        canonical = normalize_template_url(PROFILES[checklist_type]["url"])
        supplied = normalize_template_url(source)
    except ValueError as exc:
        sys.exit(f"Invalid template URL: {exc}")
    expected = PROFILES[checklist_type]["sha256"] if supplied == canonical else None
    return download_template(source, checklist_type, output_dir, expected_sha256=expected)


def build(base_workbook, out_path, buckets, meta=None, mode="begin", blank_unverified=True):
    meta = meta or {}
    shutil.copy(base_workbook, out_path)
    wb = openpyxl.load_workbook(out_path)

    missing_tabs = [TABS[k]["tab"] for k in TABS if TABS[k]["tab"] not in wb.sheetnames]
    if missing_tabs:
        sys.exit(f"Template is missing expected tabs: {missing_tabs}")

    ws = wb[CHECK_TAB]
    rows = find_check_rows(ws)
    if len(rows) != len(TABS):
        sys.exit(f"Could not locate all checklist rows; found {sorted(rows)}")

    # The template ships '✔' pre-filled on every item. Left alone, the deliverable
    # would claim a clean pass on checks nobody ran.
    if blank_unverified:
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, CHECK_COL).value in ("✔", "✖", "?"):
                ws.cell(r, CHECK_COL).value = None

    results = []
    changes = {}
    for key, spec in TABS.items():
        tab = spec["tab"]
        rowfn = spec["rowfn"]
        machine_cols = spec["machine_cols"]
        preserve_cols = spec["preserve_cols"]
        key_cols = spec["key_cols"]
        data = buckets.get(key, [])
        sheet = wb[tab]
        max_col = max((sheet.max_column, machine_cols, *preserve_cols))
        style = [copy.copy(sheet.cell(DATA_ROW, c)._style) for c in range(1, max_col + 1)]

        previous_keys = []
        row_images = collections.defaultdict(list)
        data_images = []
        for image in list(getattr(sheet, "_images", [])):
            position = _image_position(image)
            if position and position[0] >= DATA_ROW:
                row_images[position[0]].append((position[1], image))
                data_images.append(image)

        carry_forward = collections.defaultdict(collections.deque)
        for row in range(DATA_ROW, sheet.max_row + 1):
            machine = [sheet.cell(row, c).value for c in range(1, machine_cols + 1)]
            if not any(v not in (None, "") for v in machine):
                continue
            issue_key = _issue_key(machine, key_cols)
            previous_keys.append(issue_key)
            preserved_cells = tuple(_capture_cell(sheet.cell(row, c)) for c in preserve_cols)
            carry_forward[issue_key].append((preserved_cells, row_images.get(row, [])))

        if data_images:
            sheet._images = [image for image in sheet._images if image not in data_images]

        # Both modes replace the factual issue rows. Review mode first captured annotations
        # above; Begin mode also clears any accidental example data in a blank template.
        for row in sheet.iter_rows(min_row=DATA_ROW, max_row=sheet.max_row, max_col=max_col):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.hyperlink = None
                    cell.comment = None

        current_keys = []
        for i, rec in enumerate(data):
            values = rowfn(rec)
            issue_key = _issue_key(values, key_cols)
            current_keys.append(issue_key)
            for c, val in enumerate(values, start=1):
                cell = sheet.cell(DATA_ROW + i, c)
                cell.value = val
                cell._style = copy.copy(style[c - 1])
            if mode == "review" and carry_forward[issue_key]:
                carried_cells, carried_images = carry_forward[issue_key].popleft()
                for c, captured in zip(preserve_cols, carried_cells):
                    cell = sheet.cell(DATA_ROW + i, c)
                    _restore_cell(cell, captured)
                    cell._style = copy.copy(style[c - 1])
                for image_col, image in carried_images:
                    image.anchor = f"{get_column_letter(image_col)}{DATA_ROW + i}"
                    sheet.add_image(image)

        if mode == "review":
            changes[key] = _change_counts(previous_keys, current_keys)
        if data:
            sheet.sheet_state = "visible"
        # Image items can't be judged clean when the evidence isn't there:
        #   - no images captured at all -> both 20.1 and 20.2 unknown
        #   - images captured but no size data (external CDN) -> 20.2 can't confirm "no
        #     oversize", so a 0 result is unknown, not clean (20.1 is still valid)
        no_images = key in ("20.1", "20.2") and not meta.get("images_crawled", True)
        no_sizes = (key == "20.2" and not data
                    and meta.get("images_crawled", True)
                    and not meta.get("images_sizes_known", True))
        if no_images or no_sizes:
            mark = "?"
        else:
            mark = "✖" if data else "✔"
        ws.cell(rows[key], CHECK_COL).value = mark
        results.append((key, tab, len(data), mark))

    wb.save(out_path)
    return results, changes


# --------------------------------------------------------------------------- main

def default_template(checklist_type):
    """Return the canonical bundled blank template for a checklist family."""
    d = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "template")
    return os.path.join(d, PROFILES[checklist_type]["template"])


def main():
    p = argparse.ArgumentParser(description="Onsite checklist extraction (items 17.1-20.2)")
    p.add_argument("--url", help="Start URL to crawl in Spider mode")
    p.add_argument("--sitemap", help="Sitemap URL for route selection only. The local headless "
                                      "route cannot inject it into Crawl These Sitemaps; use a "
                                      "saved crawl or MCP for a full sitemap audit.")
    p.add_argument("--client", required=True, help="Client name, used in the output filename")
    p.add_argument("--checklist-type", choices=sorted(PROFILES),
                   help="Checklist family: seo or seo_geo. Auto-detected from an attached base workbook.")
    p.add_argument("--mode", choices=("begin", "review"), default="begin",
                   help="begin starts from a blank template; review updates a previous completed onsite")
    p.add_argument("--template", help="Blank checklist .xlsx for Begin mode")
    p.add_argument("--template-url", help="Approved FirstPage GitHub template URL for Begin mode")
    p.add_argument("--previous-workbook", help="Previous completed onsite .xlsx for Review mode")
    p.add_argument("--out-dir", default=".", help="Where to write the deliverable")
    p.add_argument("--config", help="Path to a .seospiderconfig (else SF defaults)")
    p.add_argument("--crawl-file", help="Export from a saved .dbseospider crawl instead of "
                                        "crawling. Uses the settings that crawl was run with.")
    p.add_argument("--exports-dir", help="Reuse existing CSV exports; skips crawling")
    p.add_argument("--date", help="YYYYMMDD for the filename (default: today)")
    p.add_argument("--keep-template-ticks", action="store_true",
                   help="Do NOT blank the template's default ticks on unverified items")
    a = p.parse_args()

    if not (a.exports_dir or a.crawl_file or a.url or a.sitemap):
        p.error("give --url, --crawl-file, or --exports-dir")
    if a.sitemap and not (a.crawl_file or a.exports_dir):
        p.error("the local headless route does not support sitemap audits. Use a saved crawl "
                "configured with Crawl These Sitemaps, or use the Screaming Frog MCP route")

    if a.mode == "review" and not a.previous_workbook:
        p.error("Review mode requires --previous-workbook")
    if a.mode == "begin" and a.previous_workbook:
        p.error("--previous-workbook is only valid in Review mode")
    if a.mode == "review" and a.template:
        p.error("Review mode uses --previous-workbook as its base; do not also pass --template")
    if a.mode == "review" and a.template_url:
        p.error("Review mode uses --previous-workbook; a blank --template-url is not valid")
    if a.mode == "begin" and not (a.template or a.checklist_type):
        p.error("Begin mode requires --checklist-type when --template is not supplied")

    template_tmp = None
    if a.mode == "review":
        base_workbook = a.previous_workbook
    elif a.template:
        base_workbook = a.template
    else:
        template_tmp = tempfile.TemporaryDirectory(prefix="onsite-template-")
        base_workbook = resolve_begin_template(
            None, a.template_url, a.checklist_type, template_tmp.name,
        )
    if not os.path.exists(base_workbook):
        sys.exit(f"Base workbook not found: {base_workbook}")

    detected_type = detect_profile(base_workbook)
    if a.checklist_type and a.checklist_type != detected_type:
        sys.exit(f"Checklist type mismatch: selected {a.checklist_type}, but the workbook is {detected_type}")
    checklist_type = a.checklist_type or detected_type

    date = a.date or datetime.date.today().strftime("%Y%m%d")

    if a.exports_dir:
        exports = os.path.abspath(a.exports_dir)
        print(f"[reuse] exports <- {exports}")
    else:
        exports = run_crawl(a.url, a.sitemap,
                            os.path.join(os.path.abspath(a.out_dir), "_sf_exports"),
                            a.config, client=a.client, date=date, crawl_file=a.crawl_file)

    buckets, meta = build_buckets(exports)
    os.makedirs(os.path.abspath(a.out_dir), exist_ok=True)
    filename = PROFILES[checklist_type]["output"].format(client=a.client, date=date)
    out_path = os.path.join(os.path.abspath(a.out_dir), filename)
    if os.path.abspath(base_workbook) == os.path.abspath(out_path):
        sys.exit("Refusing to overwrite the source workbook. Choose a different output name, "
                 "date, or directory.")
    results, changes = build(
        base_workbook, out_path, buckets, meta, mode=a.mode,
        blank_unverified=(a.mode == "begin" and not a.keep_template_ticks),
    )

    print()
    print(f"{'ITEM':<36}{'ROWS':>6}   MARK")
    print("-" * 52)
    for key, tab, n, mark in results:
        print(f"{tab:<36}{n:>6}   {mark}")
    if a.mode == "review":
        print()
        print(f"{'ITEM':<8}{'NEW':>6}{'PERSIST':>10}{'RESOLVED':>10}")
        print("-" * 34)
        for key in TABS:
            delta = changes[key]
            print(f"{key:<8}{delta['new']:>6}{delta['persistent']:>10}{delta['resolved']:>10}")
    print()
    print("WROTE:", out_path)
    print(f"thresholds: title>{TITLE_PX}px  desc>{DESC_PX}px  image>{IMG_KB}kB")
    if not meta.get("images_crawled", True):
        print()
        print("!! WARNING: the crawl captured NO images. Items 20.1/20.2 are marked '?' "
              "(unknown), NOT clean.\n"
              "   Likely image crawling was off, or JS rendering was off on a site that "
              "loads images via\n"
              "   JavaScript. Re-crawl with images enabled to audit 20.1/20.2.")
    elif not meta.get("images_sizes_known", True):
        print()
        print("!! NOTE: images were found but none had size data - they are served from an "
              "external CDN\n"
              "   (e.g. Prismic/Cloudinary), which SF lists outside Images:All. 20.1 (missing "
              "alt) is valid,\n"
              "   but 20.2 (oversize) can't be judged and is marked '?'. To size external "
              "images, enable\n"
              "   'Check Images' / external resource crawling and re-crawl.")
    if template_tmp:
        template_tmp.cleanup()


if __name__ == "__main__":
    main()
